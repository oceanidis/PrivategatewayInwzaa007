import pandas as pd
from openpyxl import Workbook
import time
from threading import Event

from privategateway import agent_plugin
from privategateway.import_pipeline import StreamingSanitizationSession
from privategateway.key_provider import init_project


def test_streaming_session_uses_one_mapping_for_multiple_batches(tmp_path):
    policy = tmp_path / "policy.yaml"
    policy.write_text(
        "\n".join([
            "security:",
            "  require_presidio: false",
            "columns:",
            "  customer_id: tokenize",
            "default:",
            "  unknown_column: review_required",
        ]),
        encoding="utf-8",
    )
    key_root = tmp_path / "keys"
    secure_root = tmp_path / "secure"
    init_project("analytics", key_root=key_root)

    session = StreamingSanitizationSession(
        policy_path=policy,
        project_id="analytics",
        job_id="one_job",
        raw_payload=b"workbook-placeholder",
        input_type="excel",
        secure_root=secure_root,
        key_root=key_root,
        scan_mode="sealed_analytics",
    )
    first = session.sanitize_frame(pd.DataFrame({"customer_id": ["C-001"]}))
    second = session.sanitize_frame(pd.DataFrame({"customer_id": ["C-001"]}))
    result = session.finalize()

    assert first.safe_dataset.iloc[0, 0] == second.safe_dataset.iloc[0, 0]
    assert result.mapping_table is not None
    assert result.can_export


def test_excel_export_uses_one_session_across_batches(tmp_path, monkeypatch):
    source = tmp_path / "input.xlsx"
    target = tmp_path / "safe.xlsx"
    policy = tmp_path / "policy.yaml"
    policy.write_text(
        "\n".join([
            "security:",
            "  require_presidio: false",
            "columns:",
            "  customer_id: tokenize",
            "default:",
            "  unknown_column: review_required",
        ]),
        encoding="utf-8",
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["customer_id"])
    sheet.append(["C-001"])
    sheet.append(["C-002"])
    workbook.save(source)
    monkeypatch.setattr(agent_plugin, "GATEWAY_ROOT", tmp_path / "gateway")
    monkeypatch.setattr(agent_plugin, "KEY_ROOT", tmp_path / "gateway" / ".privacy_gateway" / "keys")
    monkeypatch.setattr(agent_plugin, "EXCEL_BATCH_ROWS", 1)

    exported = agent_plugin.sanitize_local_file_to_file(
        str(source),
        str(target),
        input_type="excel",
        project_id="analytics",
        policy_path=str(policy),
        auto_policy=False,
        scan_mode="sealed_analytics",
    )

    assert target.exists()
    assert exported["redaction_report"]["action_counts"]["tokenize"] == 2
    assert exported["redaction_report"]["job_id"].startswith("mcp_export_")


def test_excel_export_writes_synthesized_missing_numeric_values_as_blank_cells(tmp_path, monkeypatch):
    source = tmp_path / "input.xlsx"
    target = tmp_path / "safe.xlsx"
    policy = tmp_path / "policy.yaml"
    policy.write_text(
        "\n".join([
            "security:",
            "  require_presidio: false",
            "columns:",
            "  amount: synthesize",
            "default:",
            "  unknown_column: review_required",
        ]),
        encoding="utf-8",
    )
    workbook = Workbook()
    workbook.active.append(["amount"])
    workbook.active.append([100.0])
    workbook.active.append([None])
    workbook.save(source)
    monkeypatch.setattr(agent_plugin, "GATEWAY_ROOT", tmp_path / "gateway")
    monkeypatch.setattr(agent_plugin, "KEY_ROOT", tmp_path / "gateway" / ".privacy_gateway" / "keys")

    agent_plugin.sanitize_local_file_to_file(
        str(source), str(target), "excel", "analytics", str(policy), False, "sealed_analytics"
    )

    assert target.exists()


def test_large_export_returns_background_job_and_exposes_safe_completion(tmp_path, monkeypatch):
    source = tmp_path / "input.xlsx"
    target = tmp_path / "safe.xlsx"
    policy = tmp_path / "policy.yaml"
    policy.write_text(
        "\n".join([
            "security:",
            "  require_presidio: false",
            "columns:",
            "  customer_id: tokenize",
            "default:",
            "  unknown_column: review_required",
        ]),
        encoding="utf-8",
    )
    workbook = Workbook()
    workbook.active.append(["customer_id"])
    workbook.active.append(["C-001"])
    workbook.save(source)
    monkeypatch.setattr(agent_plugin, "GATEWAY_ROOT", tmp_path / "gateway")
    monkeypatch.setattr(agent_plugin, "KEY_ROOT", tmp_path / "gateway" / ".privacy_gateway" / "keys")
    monkeypatch.setattr(agent_plugin, "ASYNC_EXPORT_SIZE_BYTES", 1)

    started = agent_plugin.sanitize_local_file_to_file(
        str(source), str(target), "excel", "analytics", str(policy), False, "sealed_analytics"
    )

    assert started["status"] in {"queued", "running"}
    for _ in range(50):
        status = agent_plugin.get_export_job_status(started["job_id"])
        if status["status"] in {"completed", "failed"}:
            break
        time.sleep(0.05)
    assert status["status"] == "completed"
    assert status["result"]["redaction_report"]["action_counts"]["tokenize"] == 1


def test_large_export_reuses_running_job_for_identical_request(tmp_path, monkeypatch):
    monkeypatch.setattr(agent_plugin, "_ensure_project", lambda project_id: project_id or "default")
    source = tmp_path / "input.xlsx"
    target = tmp_path / "safe.xlsx"
    workbook = Workbook()
    workbook.active.append(["customer_id"])
    workbook.active.append(["C-001"])
    workbook.save(source)
    monkeypatch.setattr(agent_plugin, "ASYNC_EXPORT_SIZE_BYTES", 1)
    started = Event()
    release = Event()
    calls = []

    def slow_export(*args):
        calls.append(args)
        started.set()
        release.wait(timeout=5)
        return {"output_path": str(target), "redaction_report": {}}

    monkeypatch.setattr(agent_plugin, "_sanitize_local_file_to_file_sync", slow_export)

    first = agent_plugin.sanitize_local_file_to_file(
        str(source), str(target), "excel", "analytics", None, True, "sealed_analytics"
    )
    assert started.wait(timeout=2)
    second = agent_plugin.sanitize_local_file_to_file(
        str(source), str(target), "excel", "analytics", None, True, "sealed_analytics"
    )
    release.set()

    assert second["job_id"] == first["job_id"]
    assert len(calls) == 1


def test_export_job_status_includes_safe_progress_fields():
    job_id = "export_progress"
    agent_plugin._EXPORT_JOBS[job_id] = agent_plugin.ExportJob(
        status="running",
        stage="transform",
        rows_processed=12000,
        sheets_processed=2,
    )
    try:
        status = agent_plugin.get_export_job_status(job_id)
    finally:
        agent_plugin._EXPORT_JOBS.pop(job_id, None)

    assert status == {
        "job_id": job_id,
        "status": "running",
        "stage": "transform",
        "rows_processed": 12000,
        "sheets_processed": 2,
    }


def test_auto_policy_reads_excel_preview_once_before_streaming(tmp_path, monkeypatch):
    source = tmp_path / "input.xlsx"
    target = tmp_path / "safe.xlsx"
    workbook = Workbook()
    workbook.active.append(["customer_id", "status"])
    workbook.active.append(["C-001", "ACTIVE"])
    workbook.save(source)
    monkeypatch.setattr(agent_plugin, "GATEWAY_ROOT", tmp_path / "gateway")
    monkeypatch.setattr(agent_plugin, "KEY_ROOT", tmp_path / "gateway" / ".privacy_gateway" / "keys")
    original = agent_plugin.read_preview_payloads
    calls = []

    def counted(*args, **kwargs):
        calls.append((args, kwargs))
        return original(*args, **kwargs)

    monkeypatch.setattr(agent_plugin, "read_preview_payloads", counted)

    agent_plugin.sanitize_local_file_to_file(
        str(source), str(target), "excel", "analytics", None, True, "sealed_analytics"
    )

    assert len(calls) == 1


def test_background_export_exposes_exception_type_without_exception_message(tmp_path, monkeypatch):
    monkeypatch.setattr(
        agent_plugin,
        "_sanitize_local_file_to_file_sync",
        lambda *args: (_ for _ in ()).throw(PermissionError("raw value must stay private")),
    )

    started = agent_plugin._start_export_job(
        tmp_path / "input.xlsx",
        tmp_path / "safe.xlsx",
        "excel",
        "analytics",
        None,
        True,
        "sealed_analytics",
    )

    for _ in range(50):
        status = agent_plugin.get_export_job_status(started["job_id"])
        if status["status"] == "failed":
            break
        time.sleep(0.01)
    assert status == {
        "job_id": started["job_id"],
        "status": "failed",
        "error_code": "export_failed",
        "error_type": "PermissionError",
        "error_stage": "export",
    }
