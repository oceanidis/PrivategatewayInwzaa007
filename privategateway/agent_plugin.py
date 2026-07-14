from __future__ import annotations

import hashlib
import json
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from threading import Lock, Thread
from time import monotonic
from typing import Any, Callable
from uuid import uuid4

import pandas as pd
import yaml

from .file_inputs import FilePayload, normalize_file_type, read_payloads, read_preview_payloads
from .import_pipeline import FrameSanitizationFailure, StreamingSanitizationSession, sanitize_import
from .key_provider import init_project
from .masker import GATEWAY_ROOT, handle_request
from .policy_generator import infer_policy, infer_token_domains


DEFAULT_PROJECT_ID = os.environ.get("PRIVACY_GATEWAY_PROJECT_ID", "default")
KEY_ROOT = GATEWAY_ROOT / ".privacy_gateway" / "keys"
EXCEL_BATCH_ROWS = 10_000
ASYNC_EXPORT_SIZE_BYTES = 25 * 1024 * 1024
_EXPORT_JOBS: dict[str, "ExportJob"] = {}
_EXPORT_JOBS_BY_REQUEST: dict[str, str] = {}
_EXPORT_JOBS_LOCK = Lock()


@dataclass
class ExportJob:
    status: str = "queued"
    result: dict[str, Any] | None = None
    error_code: str | None = None
    error_type: str | None = None
    error_stage: str | None = None
    stage: str = "queued"
    rows_processed: int = 0
    sheets_processed: int = 0
    started_at: float | None = None
    completed_at: float | None = None
    request_key: str | None = None


class ExportFailure(RuntimeError):
    def __init__(self, stage: str, cause: Exception) -> None:
        super().__init__(stage)
        self.stage = stage
        self.cause = cause


def sanitize_file(
    path: str,
    input_type: str | None = None,
    project_id: str | None = None,
    policy_path: str | None = None,
    auto_policy: bool = True,
) -> dict[str, Any]:
    source = Path(path).expanduser().resolve()
    normalized_type = normalize_file_type(source, input_type)
    project = _ensure_project(project_id)
    results = [_sanitize_payload(item, project, policy_path, auto_policy) for item in read_payloads(source, normalized_type)]
    if len(results) == 1:
        return results[0]["result"]
    return {
        "members": {item["name"]: item["result"]["safe_dataset"] for item in results},
        "redaction_report": {"members": [item["result"]["redaction_report"] for item in results]},
    }


def preview_local_file(
    path: str,
    input_type: str | None = None,
    project_id: str | None = None,
    policy_path: str | None = None,
    preview_rows: int = 10,
    auto_policy: bool = True,
) -> dict[str, Any]:
    """Return a sanitized bounded structure preview without full-file export."""
    source = Path(path).expanduser().resolve()
    normalized_type = normalize_file_type(source, input_type)
    project = _ensure_project(project_id)
    previews: list[dict[str, Any]] = []
    for payload in read_preview_payloads(source, normalized_type, preview_rows=preview_rows):
        if payload.input_type == "text":
            result = sanitize_text(str(payload.data), project, policy_path)
            previews.append({
                "name": payload.name,
                "columns": [],
                "rows": len(str(payload.data).splitlines()),
                "sample": result["safe_dataset"],
                "redaction_report": result["redaction_report"],
            })
            continue
        frame = payload.data
        assert isinstance(frame, pd.DataFrame)
        result = _sanitize_frame_direct(frame, project, policy_path, auto_policy)
        inferred = infer_policy(frame) if auto_policy and policy_path is None else None
        previews.append({
            "name": payload.name,
            "rows": int(frame.shape[0]),
            "columns": [str(column) for column in frame.columns],
            "inferred_types": {str(column): str(dtype) for column, dtype in frame.dtypes.items()},
            "sample": result.safe_dataset.to_dict(orient="records"),
            "suggested_policy": inferred[0] if inferred else None,
            "inferred_roles": {
                decision.source_name: decision.role for decision in inferred[2]
            } if inferred else None,
            "redaction_report": result.redaction_report.to_safe_dict(),
        })
    return {"preview_rows": preview_rows, "sheet_count": len(previews), "sheets": previews}


def sanitize_local_file_to_file(
    input_path: str,
    output_path: str,
    input_type: str | None = None,
    project_id: str | None = None,
    policy_path: str | None = None,
    auto_policy: bool = True,
    scan_mode: str = "fast",
) -> dict[str, Any]:
    source = Path(input_path).expanduser().resolve()
    target = Path(output_path).expanduser().resolve()
    normalized_type = normalize_file_type(source, input_type)
    project = _ensure_project(project_id)
    if source.stat().st_size >= ASYNC_EXPORT_SIZE_BYTES:
        request_key = _export_request_key(
            source, target, normalized_type, project, policy_path, auto_policy, scan_mode
        )
        return _start_export_job(
            source, target, normalized_type, project, policy_path, auto_policy, scan_mode, request_key
        )
    return _sanitize_local_file_to_file_sync(
        source, target, normalized_type, project, policy_path, auto_policy, scan_mode
    )


def get_export_job_status(job_id: str) -> dict[str, Any]:
    with _EXPORT_JOBS_LOCK:
        job = _EXPORT_JOBS.get(job_id)
        if job is None:
            raise ValueError("export job was not found")
        response: dict[str, Any] = {"job_id": job_id, "status": job.status}
        if job.status in {"queued", "running"}:
            response.update({
                "stage": job.stage,
                "rows_processed": job.rows_processed,
                "sheets_processed": job.sheets_processed,
            })
            if job.started_at is not None:
                response["elapsed_seconds"] = round(monotonic() - job.started_at, 3)
        if job.status == "completed":
            response["result"] = job.result
        elif job.status == "failed":
            response["error_code"] = job.error_code or "export_failed"
            response["error_type"] = job.error_type or "RuntimeError"
            response["error_stage"] = job.error_stage or "export"
        return response


def _export_request_key(
    source: Path,
    target: Path,
    normalized_type: str,
    project: str,
    policy_path: str | None,
    auto_policy: bool,
    scan_mode: str,
) -> str:
    stat = source.stat()
    payload = {
        "auto_policy": auto_policy,
        "input_type": normalized_type,
        "policy_path": str(Path(policy_path).expanduser().resolve()) if policy_path else None,
        "project": project,
        "scan_mode": scan_mode,
        "source": str(source),
        "source_mtime_ns": stat.st_mtime_ns,
        "source_size": stat.st_size,
        "target": str(target),
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()


def _start_export_job(
    source: Path,
    target: Path,
    normalized_type: str,
    project: str,
    policy_path: str | None,
    auto_policy: bool,
    scan_mode: str,
    request_key: str | None = None,
) -> dict[str, Any]:
    request_key = request_key or f"manual_{uuid4().hex}"
    with _EXPORT_JOBS_LOCK:
        active_job_id = _EXPORT_JOBS_BY_REQUEST.get(request_key)
        active_job = _EXPORT_JOBS.get(active_job_id) if active_job_id else None
        if active_job is not None and active_job.status in {"queued", "running"}:
            return {"job_id": active_job_id, "status": active_job.status, "reused": True}
        job_id = f"export_{uuid4().hex}"
        job = ExportJob(request_key=request_key)
        _EXPORT_JOBS[job_id] = job
        _EXPORT_JOBS_BY_REQUEST[request_key] = job_id

    def progress(stage: str, rows_processed: int | None = None, sheets_processed: int | None = None) -> None:
        with _EXPORT_JOBS_LOCK:
            job.stage = stage
            if rows_processed is not None:
                job.rows_processed = rows_processed
            if sheets_processed is not None:
                job.sheets_processed = sheets_processed

    def run() -> None:
        with _EXPORT_JOBS_LOCK:
            job.status = "running"
            job.stage = "prepare"
            job.started_at = monotonic()
        try:
            result = _sanitize_local_file_to_file_sync(
                source, target, normalized_type, project, policy_path, auto_policy, scan_mode, progress
            )
        except Exception as exc:
            with _EXPORT_JOBS_LOCK:
                job.status = "failed"
                job.stage = "failed"
                job.completed_at = monotonic()
                job.error_code = "export_failed"
                job.error_type = type(exc.cause).__name__ if isinstance(exc, ExportFailure) else type(exc).__name__
                job.error_stage = exc.stage if isinstance(exc, ExportFailure) else "export"
        else:
            with _EXPORT_JOBS_LOCK:
                job.status = "completed"
                job.stage = "completed"
                job.completed_at = monotonic()
                job.result = result

    Thread(target=run, name=f"privategateway-{job_id}", daemon=True).start()
    return {"job_id": job_id, "status": "queued"}


def _sanitize_local_file_to_file_sync(
    source: Path,
    target: Path,
    normalized_type: str,
    project: str,
    policy_path: str | None,
    auto_policy: bool,
    scan_mode: str,
    progress_callback: Callable[[str, int | None, int | None], None] | None = None,
) -> dict[str, Any]:
    progress = progress_callback or (lambda stage, rows_processed=None, sheets_processed=None: None)
    if normalized_type == "excel":
        generated: Path | None = None
        try:
            if auto_policy and policy_path is None:
                progress("policy")
                generated, review_columns = _prepare_workbook_policy(source)
                if review_columns:
                    raise PermissionError(
                        "privacy gateway blocked export before full scan: review_required columns: "
                        + ", ".join(review_columns)
                    )
            return _stream_sanitize_excel(
                source, target, project, generated or policy_path, scan_mode, progress
            )
        finally:
            if generated:
                generated.unlink(missing_ok=True)

    progress("sanitize")
    exported = sanitize_file(str(source), normalized_type, project, policy_path, auto_policy)
    progress("write")
    target.parent.mkdir(parents=True, exist_ok=True)
    safe = exported["safe_dataset"]
    frame = pd.DataFrame(safe if isinstance(safe, list) else [safe])
    suffix = target.suffix.lower()
    if suffix in {".csv", ".tsv", ".psv"}:
        frame.to_csv(target, index=False, sep={".csv": ",", ".tsv": "\t", ".psv": "|"}[suffix])
    elif suffix in {".xlsx", ".xlsm"}:
        frame.to_excel(target, index=False)
    else:
        target.write_text(json.dumps(safe, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "output_path": str(target),
        "metadata": {"output_path": str(target), "sha256": _sha256_file(target), "rows": len(frame), "columns": len(frame.columns)},
        "redaction_report": exported["redaction_report"],
    }


def _stream_sanitize_excel(
    source: Path,
    target: Path,
    project: str,
    policy_path: str | Path | None,
    scan_mode: str,
    progress: Callable[[str, int | None, int | None], None],
) -> dict[str, Any]:
    """Sanitize every sheet in bounded row batches and write a data-only workbook."""
    from openpyxl import load_workbook
    from xlsxwriter import Workbook

    if policy_path is None:
        raise ValueError("policy_path is required for Excel export")
    target.parent.mkdir(parents=True, exist_ok=True)
    partial_target = target.with_name(target.name + ".partial")
    partial_target.unlink(missing_ok=True)
    progress("open_workbook", 0, 0)
    reader = load_workbook(source, read_only=True, data_only=True)
    writer = Workbook(str(partial_target), {"constant_memory": True})
    writer_closed = False
    metadata: dict[str, dict[str, int]] = {}
    total_rows = 0
    stage = "secure_store"
    try:
        progress(stage, 0, 0)
        session = StreamingSanitizationSession(
            policy_path=policy_path,
            project_id=project,
            job_id=f"mcp_export_{uuid4().hex}",
            raw_payload=source.read_bytes(),
            input_type="excel",
            secure_root=GATEWAY_ROOT / ".privacy_gateway" / "secure",
            key_root=KEY_ROOT,
            scan_mode=scan_mode,
        )
        for worksheet in reader.worksheets:
            output_sheet = writer.add_worksheet(_safe_sheet_name(worksheet.title))
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                metadata[worksheet.title] = {"rows": 0, "columns": 0}
                progress("transform", total_rows, len(metadata))
                continue
            columns = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(header)]
            batch: list[list[Any]] = []
            output_columns: list[str] | None = None
            sheet_rows = 0
            for row in rows:
                batch.append(list(row[: len(columns)]))
                if len(batch) < EXCEL_BATCH_ROWS:
                    continue
                stage = "transform"
                sheet_rows, output_columns = _stream_batch(batch, columns, output_sheet, session, sheet_rows, output_columns)
                total_rows += len(batch)
                batch = []
                progress(stage, total_rows, len(metadata))
            if batch:
                stage = "transform"
                sheet_rows, output_columns = _stream_batch(batch, columns, output_sheet, session, sheet_rows, output_columns)
                total_rows += len(batch)
            if output_columns is None:
                output_sheet.write_row(0, 0, columns)
                output_columns = columns
            metadata[worksheet.title] = {"rows": sheet_rows, "columns": len(output_columns)}
            progress("transform", total_rows, len(metadata))

        stage = "finalize"
        progress(stage, total_rows, len(metadata))
        final_result = session.finalize()
        if not final_result.can_export:
            reasons = final_result.redaction_report.block_reasons or ["privacy review required"]
            raise PermissionError("privacy gateway blocked export: " + ", ".join(reasons))
        stage = "write"
        progress(stage, total_rows, len(metadata))
        writer.close()
        writer_closed = True
        partial_target.replace(target)
    except ExportFailure:
        raise
    except Exception as exc:
        raise ExportFailure(stage, exc) from exc
    finally:
        if not writer_closed:
            try:
                writer.close()
            except Exception:
                pass
        reader.close()
        partial_target.unlink(missing_ok=True)
    return {
        "output_path": str(target),
        "metadata": {"output_path": str(target), "sha256": _sha256_file(target), "sheet_count": len(metadata), "sheets": metadata},
        "redaction_report": final_result.redaction_report.to_safe_dict(),
    }


def _stream_batch(
    batch: list[list[Any]],
    columns: list[str],
    output_sheet: Any,
    session: StreamingSanitizationSession,
    row_count: int,
    output_columns: list[str] | None,
) -> tuple[int, list[str] | None]:
    frame = pd.DataFrame(batch, columns=columns)
    try:
        result = session.sanitize_frame(frame)
    except Exception as exc:
        if isinstance(exc, FrameSanitizationFailure):
            raise ExportFailure(f"transform:{exc.action}", exc.cause) from exc
        raise ExportFailure("transform", exc) from exc
    if not result.can_export:
        reasons = result.redaction_report.block_reasons or ["privacy review required"]
        raise PermissionError("privacy gateway blocked export: " + ", ".join(reasons))
    safe = result.safe_dataset
    if output_columns is None:
        output_columns = [str(column) for column in safe.columns]
        output_sheet.write_row(0, 0, output_columns)
    try:
        for row_offset, values in enumerate(safe.itertuples(index=False, name=None), start=row_count + 1):
            output_sheet.write_row(row_offset, 0, [None if pd.isna(value) else value for value in values])
    except Exception as exc:
        raise ExportFailure("append", exc) from exc
    return row_count + len(safe), output_columns


def _prepare_workbook_policy(source: Path) -> tuple[Path, list[str]]:
    frames = [payload.data for payload in read_preview_payloads(source, "excel", preview_rows=100) if payload.input_type == "dataframe"]
    frame = pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()
    _, _, decisions, _ = infer_policy(frame)
    review_columns = list(dict.fromkeys(decision.source_name for decision in decisions if decision.action == "review_required"))
    return _policy_for_frame(frame), review_columns


def _sanitize_frame_direct(
    frame: pd.DataFrame,
    project: str,
    policy_path: str | None,
    auto_policy: bool,
    scan_mode: str = "strict",
):
    generated = _policy_for_frame(frame) if auto_policy and policy_path is None else None
    try:
        return sanitize_import(
            input_data=frame,
            input_type="dataframe",
            policy_path=generated or policy_path,
            project_id=project,
            job_id=f"mcp_{uuid4().hex}",
            secure_root=GATEWAY_ROOT / ".privacy_gateway" / "secure",
            key_root=KEY_ROOT,
            scan_mode=scan_mode,
        )
    finally:
        if generated:
            generated.unlink(missing_ok=True)


def sanitize_text(text: str, project_id: str | None = None, policy_path: str | None = None) -> dict[str, Any]:
    project = _ensure_project(project_id)
    return _exportable_response(handle_request({"input_type": "text", "data": text, "project_id": project, "policy_path": policy_path}))


def sanitize_records(records: list[dict[str, Any]], project_id: str | None = None, policy_path: str | None = None) -> dict[str, Any]:
    project = _ensure_project(project_id)
    return _exportable_response(handle_request({"input_type": "dataframe", "data": records, "project_id": project, "policy_path": policy_path}))


def _sanitize_payload(payload: FilePayload, project: str, policy_path: str | None, auto_policy: bool) -> dict[str, Any]:
    if payload.input_type == "text":
        return {"name": payload.name, "result": sanitize_text(str(payload.data), project, policy_path)}
    frame = payload.data
    assert isinstance(frame, pd.DataFrame)
    result = _sanitize_frame_direct(frame, project, policy_path, auto_policy)
    return {
        "name": payload.name,
        "result": {
            "safe_dataset": result.safe_dataset.to_dict(orient="records"),
            "redaction_report": result.redaction_report.to_safe_dict(),
        },
    }


def _policy_for_frame(frame: pd.DataFrame) -> Path:
    columns, buckets, decisions, subject = infer_policy(frame)
    payload = {
        "security": {"require_presidio": True, "raw_ttl_hours": 24, "mapping_ttl_days": 30, "reject_duplicate_job_id": True},
        "date_shift": {"scope": "subject", "subject_column": subject or "customer_id", "min_days": 1, "max_days": 30, "direction": "both", "stability": "project"},
        "time_shift": {"scope": "subject", "subject_column": subject or "customer_id", "min_minutes": 1, "max_minutes": 720, "direction": "both", "stability": "project"},
        "columns": columns,
        "token_domains": infer_token_domains(decisions),
        "default": {"unknown_column": "review_required"},
        "bucket": buckets,
        "custom_recognizers": [],
    }
    handle = tempfile.NamedTemporaryFile(mode="w", suffix=".yaml", encoding="utf-8", delete=False)
    with handle:
        yaml.safe_dump(payload, handle, allow_unicode=True, sort_keys=False)
    return Path(handle.name)


def _ensure_project(project_id: str | None) -> str:
    project = project_id or DEFAULT_PROJECT_ID
    init_project(project, key_root=KEY_ROOT)
    return project


def _exportable_response(response: dict[str, Any]) -> dict[str, Any]:
    if not response.get("ok"):
        raise RuntimeError("PrivateGateway rejected the import")
    if not response.get("can_export"):
        raise PermissionError("PrivateGateway blocked downstream export")
    return {"safe_dataset": response["safe_dataset"], "redaction_report": response["redaction_report"]}


def _safe_sheet_name(name: str) -> str:
    value = "".join("_" if character in "[]:*?/\\" else character for character in str(name)).strip()
    return (value or "Sheet1")[:31]


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
