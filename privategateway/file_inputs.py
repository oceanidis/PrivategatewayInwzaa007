from __future__ import annotations

import gzip
import json
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal
from zipfile import ZIP_STORED, ZipFile

import pandas as pd
import yaml


_ALIASES = {
    "csv": "csv", "tsv": "tsv", "psv": "psv", "excel": "excel",
    "xlsx": "excel", "xls": "excel", "xlsm": "excel", "parquet": "parquet",
    "feather": "feather", "orc": "orc", "json": "json", "jsonl": "jsonl",
    "ndjson": "jsonl", "xml": "xml", "yaml": "yaml", "yml": "yaml",
    "txt": "text", "text": "text", "log": "text", "zip": "zip",
    "gz": "gzip", "gzip": "gzip",
}
_ARCHIVES = {"zip", "gzip"}


@dataclass(frozen=True)
class FilePayload:
    name: str
    input_type: Literal["dataframe", "text"]
    data: pd.DataFrame | str


def normalize_file_type(path: Path, input_type: str | None = None) -> str:
    value = (input_type or path.suffix.lstrip(".")).strip().lower()
    try:
        return _ALIASES[value]
    except KeyError as exc:
        raise ValueError(f"unsupported file type: {value or '<none>'}") from exc


def read_payloads(
    path: Path,
    input_type: str | None = None,
    *,
    max_input_bytes: int = 100_000_000,
    max_archive_members: int = 100,
    max_extracted_bytes: int = 500_000_000,
) -> list[FilePayload]:
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError("input file was not found")
    if path.stat().st_size > max_input_bytes:
        raise ValueError("input file exceeds the maximum allowed size")
    normalized = normalize_file_type(path, input_type)
    if normalized == "zip":
        return _read_zip(path, max_archive_members, max_extracted_bytes)
    if normalized == "gzip":
        with gzip.open(path, "rb") as source:
            data = _read_limited(source, max_extracted_bytes)
        suffix = Path(path.stem).suffix or ".txt"
        return [_parse_bytes(path.stem, suffix, data)]
    parsed = _parse_bytes(path.name, path.suffix, path.read_bytes(), normalized)
    return parsed if isinstance(parsed, list) else [parsed]


def read_preview_payloads(
    path: Path,
    input_type: str | None = None,
    *,
    preview_rows: int = 10,
) -> list[FilePayload]:
    """Read only a bounded, non-archive sample for structure inspection."""
    if preview_rows < 1:
        raise ValueError("preview_rows must be at least 1")
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError("input file was not found")
    normalized = normalize_file_type(path, input_type)
    if normalized in _ARCHIVES:
        raise ValueError("preview is not supported for archives; preview the extracted data file")
    if normalized == "excel":
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            return _read_excel_preview_read_only(path, preview_rows)
        sheets = pd.read_excel(path, sheet_name=None, nrows=preview_rows)
        return [FilePayload(str(name), "dataframe", frame) for name, frame in sheets.items()]
    if normalized in {"csv", "tsv", "psv"}:
        separator = {"csv": ",", "tsv": "\t", "psv": "|"}[normalized]
        return [FilePayload(path.name, "dataframe", pd.read_csv(path, sep=separator, nrows=preview_rows))]
    if normalized == "text":
        lines = path.read_text(encoding="utf-8").splitlines()[:preview_rows]
        return [FilePayload(path.name, "text", "\n".join(lines))]
    payloads = read_payloads(path, normalized)
    return [
        FilePayload(payload.name, payload.input_type, payload.data.head(preview_rows))
        if payload.input_type == "dataframe" else payload
        for payload in payloads
    ]


def _read_excel_preview_read_only(path: Path, preview_rows: int) -> list[FilePayload]:
    """Read only header plus bounded rows without materializing the workbook."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    payloads: list[FilePayload] = []
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                payloads.append(FilePayload(worksheet.title, "dataframe", pd.DataFrame()))
                continue
            columns = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(header)]
            values = []
            for row in rows:
                values.append(list(row[: len(columns)]))
                if len(values) >= preview_rows:
                    break
            payloads.append(FilePayload(worksheet.title, "dataframe", pd.DataFrame(values, columns=columns)))
    finally:
        workbook.close()
    return payloads


def _read_zip(path: Path, max_members: int, max_bytes: int) -> list[FilePayload]:
    result: list[FilePayload] = []
    total = 0
    with ZipFile(path) as archive:
        entries = [item for item in archive.infolist() if not item.is_dir()]
        if len(entries) > max_members:
            raise ValueError("archive has too many members")
        for item in entries:
            if item.flag_bits & 0x1 or item.compress_type not in {ZIP_STORED, 8}:
                raise ValueError("archive member is encrypted or unsupported")
            member_path = Path(item.filename)
            if member_path.is_absolute() or ".." in member_path.parts:
                raise ValueError("unsafe archive member path")
            normalized = normalize_file_type(member_path)
            if normalized in _ARCHIVES:
                raise ValueError("nested archives are unsupported")
            total += item.file_size
            if total > max_bytes:
                raise ValueError("archive exceeds extracted-byte limit")
            with archive.open(item) as source:
                parsed = _parse_bytes(item.filename, member_path.suffix, _read_limited(source, max_bytes), normalized)
                result.extend(parsed if isinstance(parsed, list) else [parsed])
    if not result:
        raise ValueError("archive has no data files")
    return result


def _read_limited(source, limit: int) -> bytes:
    chunks: list[bytes] = []
    total = 0
    while chunk := source.read(min(1_048_576, limit + 1)):
        total += len(chunk)
        if total > limit:
            raise ValueError("payload exceeds extracted-byte limit")
        chunks.append(chunk)
    return b"".join(chunks)


def _parse_bytes(name: str, suffix: str, data: bytes, normalized: str | None = None) -> FilePayload | list[FilePayload]:
    kind = normalized or normalize_file_type(Path(name))
    if kind == "text":
        return FilePayload(name, "text", data.decode("utf-8"))
    if kind in {"csv", "tsv", "psv"}:
        separator = {"csv": ",", "tsv": "\t", "psv": "|"}[kind]
        return FilePayload(name, "dataframe", pd.read_csv(StringIO(data.decode("utf-8")), sep=separator))
    if kind == "excel":
        sheets = pd.read_excel(BytesIO(data), sheet_name=None)
        return [FilePayload(f"{sheet_name}", "dataframe", frame) for sheet_name, frame in sheets.items()]
    if kind == "json":
        payload = json.loads(data.decode("utf-8"))
        return FilePayload(name, "dataframe", pd.DataFrame(payload if isinstance(payload, list) else [payload]))
    if kind == "jsonl":
        return FilePayload(name, "dataframe", pd.read_json(BytesIO(data), lines=True))
    if kind == "yaml":
        payload = yaml.safe_load(data.decode("utf-8"))
        return FilePayload(name, "dataframe", pd.DataFrame(payload if isinstance(payload, list) else [payload]))
    if kind == "xml":
        return FilePayload(name, "dataframe", pd.read_xml(BytesIO(data)))
    if kind == "parquet":
        return FilePayload(name, "dataframe", pd.read_parquet(BytesIO(data)))
    if kind == "feather":
        return FilePayload(name, "dataframe", pd.read_feather(BytesIO(data)))
    if kind == "orc":
        return FilePayload(name, "dataframe", pd.read_orc(BytesIO(data)))
    raise ValueError(f"unsupported file type: {kind}")
